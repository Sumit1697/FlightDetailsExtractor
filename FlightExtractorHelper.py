from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
import openpyxl as ox
import requests


data = []


async def get_url_list(url, driver):
    driver.get(url)
    wait = WebDriverWait(driver, 10)
    urls = []

    try:
        elements = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'smp-accordion__tab')))
    except (NoSuchElementException, TimeoutException):
        print("No Element Found or Timeout Occurred")

    for element_count in range(len(elements)):
        elements[element_count].click()
        urls.append(find_last_element_and_return_link(driver, wait, element_count))

    # driver.close()
    return urls


def find_last_element_and_return_link(driver, wait, element_count):
    try:
        links = wait.until(EC.presence_of_all_elements_located(
            (By.XPATH, f'/html/body/main/div[2]/section/div/div[{element_count + 1}]/div/ul/li/a')))
        length = len(links) - 1
        return links[length].get_attribute('href')

    except (NoSuchElementException, TimeoutException):
        print("No Element Found or Timeout Occurred")


async def check_page_status(url):
    try:
        response = requests.get(url)
        return response.status_code
    except requests.RequestException as e:
        print(f"Error checking page status: {e}")
        return None


async def flight_schools_extractor(driver):
    try:
        data = []
        await go_to_school_and_extract('//*[@id="sl-school-list"]/li/a', driver)
        insert_all_data_in_excel(data)
    except Exception as e:
        print(f"Exception: {e}")


async def go_to_school_and_extract(xpath, driver):
    try:
        wait = WebDriverWait(driver, 1)
        no_of_pages = await get_number_of_page('/html/body/main/div/div[2]', wait)
        await school_scraper(wait, xpath, driver)
        is_pagination = await check_if_pagination(wait)
        if is_pagination:
            for no_of_page in range(no_of_pages):
                try:
                    next_button = wait.until(
                        EC.presence_of_element_located(
                            (By.CSS_SELECTOR, 'a[title="Next Page"].sl-pagination__link--filled')))
                    next_button.click()
                    await school_scraper(wait, xpath, driver)
                except (NoSuchElementException, TimeoutException):
                    print("No Element Found or Timeout Occurred")
        else:
            print("Go to next url")
    except Exception as e:
        print(f"Exception in go to school extract: {e}")


async def get_number_of_page(xpath, wait):
    try:
        no_of_pages = wait.until(EC.presence_of_all_elements_located((By.XPATH, xpath)))[0].text.split()
        return int(no_of_pages[3])
    except Exception as e:
        print(f"Exception in scraping no of pages: {e}")


async def school_scraper(wait, xpath, driver):
    try:
        try:
            school_names = wait.until(EC.presence_of_all_elements_located((By.XPATH, xpath)))
        except (NoSuchElementException, TimeoutException):
            print("No Element Found or Timeout Occurred")

        for school_count in range(len(school_names)):
            try:
                school_name = wait.until(EC.presence_of_all_elements_located((By.XPATH, xpath)))
            except (NoSuchElementException, TimeoutException):
                print("No Element Found or Timeout Occurred")

            try:
                school_link = school_name[school_count].get_attribute("href")
            except (NoSuchElementException, TimeoutException):
                print("No Element Found or Timeout Occurred")

            response_school = await check_page_status(school_link)
            if response_school == 200:
                # school_name[school_count].click()
                await get_all_details_of_flight(school_name, school_count, wait)
                driver.back()
    except Exception as e:
        print(f"Exception in School Scraper: {e}")


async def check_if_pagination(wait):
    try:
        next_button = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'a[title="Next Page"].sl-pagination__link--filled')))
        return True
    except TimeoutException:
        print("Next button not found within timeout.")
        return False
    except Exception as e:
        print("Exception at Pagination: ", e)
        return False


async def get_all_details_of_flight(school_name, school_count, wait):
    address = []
    categories = []
    program = []
    phone_number = ''
    email = ''
    try:
        try:
            school_name[school_count].click()
        except Exception as e:
            print(f"Not Clickable Exception: {e}")

        try:
            flight_school_name = wait.until(EC.presence_of_all_elements_located((By.ID, 'test-h1-name')))[
                0].text.strip().replace('\n', ' ')
        except (NoSuchElementException, TimeoutException):
            print("No Element Found or Timeout Occurred")

        try:
            about = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'sp-section')))[
                0].text.strip().replace('\n', ' ')
        except (NoSuchElementException, TimeoutException):
            print("No Element Found or Timeout Occurred")

        try:
            short_address = \
                wait.until(EC.presence_of_all_elements_located((By.XPATH, '/html/body/main/section[2]/div[1]/p')))[
                    0].text.strip().replace('\n', ' ')
        except (NoSuchElementException, TimeoutException):
            print("No Element Found or Timeout Occurred")

        try:
            phone_number = \
                wait.until(EC.presence_of_all_elements_located((By.XPATH, '/html/body/main/section[2]/div[2]/a[1]')))[
                    0].get_attribute('href')
        except (NoSuchElementException, TimeoutException):
            print("No Element Found or Timeout Occurred")

        try:
            email = \
                wait.until(EC.presence_of_all_elements_located((By.XPATH, '/html/body/main/section[2]/div[2]/a[2]')))[
                    0].get_attribute('href')
        except (NoSuchElementException, TimeoutException):
            print("No Element Found or Timeout Occurred")

        try:
            category = wait.until(
                EC.presence_of_all_elements_located((By.XPATH, '/html/body/main/section/div[1]/span')))
            for i in category:
                categories.append(i.text.strip())
        except (NoSuchElementException, TimeoutException):
            print("No Element Found or Timeout Occurred")

        try:
            programs = wait.until(
                EC.presence_of_all_elements_located((By.XPATH, '/html/body/main/section/div[2]/span')))
            for i in programs:
                program.append(i.text.strip())
        except (NoSuchElementException, TimeoutException):
            print("No Element Found or Timeout Occurred")

        try:
            view_all_button = wait.until(EC.presence_of_element_located((By.ID, 'sp-locations-view-all')))
            view_all_button.click()
        except (NoSuchElementException, TimeoutException):
            print("The 'View All Locations' button is not present. Proceeding with available locations.")

        try:
            full_addresses = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'sp-location__location')))
            for full_address in full_addresses:
                address.append(full_address.text.strip().replace('\n', ' '))
        except (NoSuchElementException, TimeoutException):
            print("No Element Found or Timeout Occurred")

        insert_data_into_excel(flight_school_name, short_address, about, phone_number, email, address, categories,
                               program)
        # print(f"\nSchool Name: {flight_school_name}, Short Address: {short_address}, About: {about}, Phone Number: {phone_number},"
        #       f"email: {email}, full address: {address}, category: {categories}")
    except Exception as e:
        print(f"Exception from get_all_details: {e}")


def insert_data_into_excel(flight_school_name, short_address, about, phone_number, email, address, categories, program):
    try:
        data.append([flight_school_name, short_address, about, phone_number, email, address, categories, program])
    except Exception as e:
        print(f"Exception Occurred in Insertion of data: {e}")


def insert_all_data_in_excel(data):
    try:
        workbook = ox.load_workbook('output.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = ox.Workbook()
        sheet = workbook.active
        headers = ['School Name', 'Short Address', 'About', 'Phone Number', 'Email', 'Address', 'Category',
                   'Program']
        sheet.append(headers)

    for record in data:
        # Convert list items to strings
        record = [', '.join(item) if isinstance(item, list) else item for item in record]
        sheet.append(record)

    workbook.save('output.xlsx')


def go_to_url_and_proceed_login(email, password_input, driver, url, wait):
    try:
        driver.get(url)
        try:
            login_button = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="col-top-login"]/div')))
            login_button.click()
        except NoSuchElementException:
            print(f"Login Button not Present")

        try:
            email_id = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="48:2;a"]')))
            email_id.send_keys(email)
        except NoSuchElementException:
            print(f"Email Id Element is not found: {NoSuchElementException}")

        try:
            password = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="61:2;a"]')))
            password.send_keys(password_input)
            password.send_keys(Keys.ENTER)
        except NoSuchElementException:
            print(f"Password element is not found: {NoSuchElementException}")
        return True
    except Exception as e:
        print(f"Exception in Login: {e}")
        return False


def get_nbaa_urls(driver, wait):
    urls = []
    try:
        member_directory = wait.until(
            EC.presence_of_all_elements_located((By.XPATH, '//*[@id="bs-navbar-collapse-top"]/ul'
                                                           '/li[3]/div/div/div/ul/li/a')))
        for i in member_directory:
            if i.get_attribute('href') not in urls:
                urls.append(i.get_attribute('href'))
        return urls
    except NoSuchElementException:
        print(f"Menu not Found: {NoSuchElementException}")


async def go_to_url_and_start_scraping(url, wait, driver):
    print("url:", url)
    if url == 'https://connect.nbaa.org/68/People':
        driver.get(url)
        company_or_people = url.split('/')[4]
        if company_or_people == 'People':
            try:
                hundred = wait.until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="background-header"]/div[2]/div['
                                                              '2]/section/div/div/div/div/div[2]/div['
                                                              '5]/div/div[3]/span[2]')))
                hundred.click()
            except NoSuchElementException:
                print(f"No Such Element Present for Doing 100 per page: {NoSuchElementException}")

            try:
                select_hundred = wait.until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="background-header"]/div['
                                                              '2]/div[2]/section/div/div/div/div/div['
                                                              '2]/div[5]/div/div[3]/span['
                                                              '2]/select/option[4]')))
                select_hundred.click()
            except NoSuchElementException:
                print(f"No able to find 100 in select element: {NoSuchElementException}")
            driver.get("https://connect.nbaa.org/common/default.aspx?id=68&pi=65&salt=75319&ps=100")
            await scrape_date_for_people(wait, driver)

        else:
            try:
                hundred = wait.until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="background-header"]/div[2]/div['
                                                              '2]/section/div/div/div/div/div[2]/div['
                                                              '5]/div/div[3]/span[2]')))
                hundred.click()
            except NoSuchElementException:
                print(f"No Such Element Present for Doing 100 per page: {NoSuchElementException}")

            try:
                select_hundred = wait.until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="background-header"]/div['
                                                              '2]/div[2]/section/div/div/div/div/div['
                                                              '2]/div[5]/div/div[3]/span['
                                                              '2]/select/option[4]')))
                select_hundred.click()
            except NoSuchElementException:
                print(f"No able to find 100 in select element: {NoSuchElementException}")
            # scrape_data_for_company(wait, driver)
    else:
        driver.get(url)
        # driver.get("https://connect.nbaa.org/common/default.aspx?id=68&pi=18&salt=93427&ps=100")
        await scrape_date_for_people(wait, driver)


async def scrape_date_for_people(wait, driver):
    try:
        names = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="background-header"]/div[2]/div[2]/section/div/div/div/div/div[2]/div[4]/div/div[1]/div/div/ul/div/div/ul/li/div/article/div/span/a')))
        for name in names:
            name.click()
            save_date_into_list(wait, driver)
            driver.back()

        await is_pagination_found(wait, driver)
    except Exception as e:
        current_url = driver.current_url
        print(f"Start Scraping again:{current_url}")
        await go_to_url_and_start_scraping(current_url, wait, driver)
        print(f"Error while performing Scraping Operation: {e}")


def save_date_into_list(wait, driver):
    person_name = ''
    website = ''
    title = ''
    title_2 = ''
    address = ''
    email = ''
    phone = ''
    fax = ''
    try:
        person_name = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="background-header"]/div/div[2]/div[2]/section/div[1]/div[2]/div/h1'))).text.strip().replace('\n', ' ')
        # print("person name",person_name)
    except (NoSuchElementException, TimeoutException):
        print(f"Name not Present: {NoSuchElementException}")

    try:
        website = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="background-header"]/div/div[2]/div[2]/section/div[1]/div[2]/div/div[2]/div[1]/div[1]/a')))[0].text.strip().replace('\n', ' ')
        # for website in websites:
        #     website = website.text.strip().replace('\n', ' ')
    except (NoSuchElementException, TimeoutException):
        print(f"Website Not Found: {NoSuchElementException}")

    try:
        title = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="background-header"]/div/div[2]/div[2]/section/div[1]/div[2]/div/div[2]/div[1]/div[1]/div[1]/span[1]'))).text.strip().replace('\n', ' ')
#         print("title: ", title)
    except (NoSuchElementException, TimeoutException):
        print(f"Title Not Found: {NoSuchElementException}")

    try:
        title_2 = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="background-header"]/div/div[2]/div[2]/section/div[1]/div[2]/div/div[2]/div[1]/div[1]/div[1]/span[3]'))).text.strip().replace('\n', ' ')
#         print("title2: ", title_2)
    except (NoSuchElementException, TimeoutException):
        print(f"Title 2 Not Found: {NoSuchElementException}")

    try:
        address = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="background-header"]/div/div[2]/div[2]/section/div[1]/div[2]/div/div[2]/div[1]/div[1]/div[2]/div/div[2]'))).text.strip().replace('\n', ' ')
#         print("address: ", address)
    except (NoSuchElementException, TimeoutException):
        print(f"Address Not Found: {NoSuchElementException}")

    try:
        email = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="background-header"]/div/div[2]/div[2]/section/div[1]/div[2]/div/div[2]/div[2]/div[2]/div/table/tbody/tr[1]/td[2]/div/span/span/a'))).get_attribute('href').split(':')[1]
#         print("email: ", email)
    except (NoSuchElementException, TimeoutException):
        print(f"Email Not Found: {NoSuchElementException}")

    try:
        phone = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="background-header"]/div/div[2]/div[2]/section/div[1]/div[2]/div/div[2]/div[2]/div[2]/div/table/tbody/tr[2]/td[2]/div/span/span'))).text.strip().replace('\n', ' ').replace('.','')
#         print("phone: ", phone)
    except (NoSuchElementException, TimeoutException):
        print(f"Phone Not Found: {NoSuchElementException}")

    try:
        fax = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="background-header"]/div/div[2]/div[2]/section/div[1]/div[2]/div/div[2]/div[2]/div[2]/div/table/tbody/tr[3]/td[2]/div/span/span'))).text.strip().replace('\n', ' ').replace('.','')
#         print("fax: ", fax)
    except (NoSuchElementException, TimeoutException):
        print(f"Fax Not Found: {NoSuchElementException}")

    data.append([person_name, website, title, title_2, address, email, phone, fax])

    if len(data) > 100:
        insert_all_data_in_excel_peoples(data)
        data.clear()


async def is_pagination_found(wait, driver):
    try:
        next_button = wait.until(EC.presence_of_element_located((By.XPATH,
                                                                 '//*[@id="background-header"]/div[2]/div[2]/section/div/div/div/div/div[2]/div[5]/div/div[2]/ul/li[5]/a')))
        next_button.click()
        await scrape_date_for_people(wait, driver)
    except NoSuchElementException:
        print(f"No Next Button Found")


def scrape_data_for_company(wait, driver):
    print("Company Found")


def insert_all_data_in_excel_peoples(data):
    try:
        workbook = ox.load_workbook('output.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = ox.Workbook()
        sheet = workbook.active
        headers = ['Person Name', 'Website', 'Title', 'Title 2', 'Address', 'Email', 'Phone',
                   'Fax']
        sheet.append(headers)

    for record in data:
        # Convert list items to strings
        record = [', '.join(item) if isinstance(item, list) else item for item in record]
        sheet.append(record)

    workbook.save('output.xlsx')